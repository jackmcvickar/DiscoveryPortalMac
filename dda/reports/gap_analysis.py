import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill

DB_PATH = "Data/dda.db"

def build_expected_periods(start_period="2024-04", end_period="2025-11"):
    start_year, start_month = map(int, start_period.split("-"))
    end_year, end_month = map(int, end_period.split("-"))
    periods = []
    for y in range(start_year, end_year + 1):
        for m in range(1, 13):
            if (y == start_year and m < start_month) or (y == end_year and m > end_month):
                continue
            periods.append(f"{y}-{m:02d}")
    return periods

def run_gap_analysis(start_period="2024-04", end_period="2025-11"):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    expected_periods = build_expected_periods(start_period, end_period)
    results = []

    cursor.execute("SELECT account_name, account_number, account_type, owner, status, notes FROM accounts_registry")
    registry_accounts = cursor.fetchall()

    for account_name, account_number, account_type, owner, status, notes in registry_accounts:
        acct_norm = account_name.strip().lower()
        last4 = account_number[-4:] if account_number else ""

        cursor.execute(
            "SELECT filepath, period FROM documents WHERE lower(account_id)=? OR account_number_last4=?",
            (acct_norm, last4)
        )
        docs = cursor.fetchall()

        found_periods = [row[1] for row in docs if row[1]]
        found_docs = [{"filename": row[0].split("/")[-1], "period": row[1]} for row in docs if row[1]]
        missing_periods = [p for p in expected_periods if p not in found_periods]

        completeness = (len(found_periods) / len(expected_periods)) * 100 if expected_periods else 0
        status_flag = "✅ Complete" if completeness == 100 else "⚠️ Missing" if completeness > 0 else "❌ None"

        results.append({
            "Account": account_name,
            "Owner": owner,
            "Type": account_type,
            "Status": status,
            "Expected": len(expected_periods),
            "Found": len(found_periods),
            "Missing": len(missing_periods),
            "Completeness": round(completeness, 1),
            "StatusFlag": status_flag,
            "FoundDocs": found_docs,
            "MissingDetails": missing_periods,
        })

    conn.close()
    return results, expected_periods

def export_matrix_excel(results, start_period, end_period, expected_periods):
    matrix = {}
    month_counts = {m: 0 for m in expected_periods}
    missing_rows = []

    for r in results:
        row = {}
        found_periods = {doc["period"] for doc in r["FoundDocs"]}
        for month in expected_periods:
            if month in found_periods:
                row[month] = "✅"
                month_counts[month] += 1
            else:
                row[month] = "⚠️"
        matrix[(r["Account"], r["Owner"])] = row

        if r["MissingDetails"]:
            missing_rows.append({
                "Account": r["Account"],
                "Owner": r["Owner"],
                "MissingPeriods": ", ".join(r["MissingDetails"])
            })

    df_matrix = pd.DataFrame.from_dict(matrix, orient="index")
    df_matrix.index = [f"{acct} ({owner})" for acct, owner in df_matrix.index]

    # Add summary row: % completeness per month across all accounts
    summary = {}
    total_accounts = len(results)
    for month in expected_periods:
        covered = sum(1 for r in results if month in {doc["period"] for doc in r["FoundDocs"]})
        completeness = (covered / total_accounts) * 100 if total_accounts else 0
        summary[month] = f"{round(completeness,1)}%"
    df_matrix.loc["Overall Completeness"] = pd.Series(summary)

    df_counts = pd.DataFrame([month_counts], index=["Statement Count"])
    df_missing = pd.DataFrame(missing_rows)

    excel_file = f"reports/gap_analysis_matrix_{start_period}_to_{end_period}.xlsx"
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df_matrix.to_excel(writer, sheet_name="Matrix")
        df_counts.to_excel(writer, sheet_name="Counts")
        df_missing.to_excel(writer, sheet_name="Missing Periods")

        workbook = writer.book
        ws = workbook["Matrix"]
        green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
        red = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

        for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value == "✅":
                    cell.fill = green
                elif cell.value == "⚠️":
                    cell.fill = orange
                elif cell.value == "❌":
                    cell.fill = red

    print(f"📊 Excel dashboard exported to {excel_file}")

if __name__ == "__main__":
    start = "2024-04"
    end = "2025-11"
    results, expected_periods = run_gap_analysis(start, end)

    print("\nGap Analysis Summary:\n")
    for r in results:
        print(
            f"{r['Account']} ({r['Owner']}) | Expected {r['Expected']} | Found {r['Found']} | "
            f"Missing {r['Missing']} | Completeness {r['Completeness']}% | {r['StatusFlag']}"
        )

    export_matrix_excel(results, start, end, expected_periods)

# end of script